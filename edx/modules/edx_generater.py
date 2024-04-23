## =======================================================
## Program: edX Marks Generaters (edx_generater)
## Author: Le Zhang
## Email: l652zhan@uwaterloo.ca
## Created Time: 2024-03-21
## Modified by:
##   [2024-03-21] - Le Zhang - CS136 (Winter 2024)
## Company: University of Waterloo
## Department: School of Computer Science
## =======================================================

import csv
import getpass
import os
import pandas as pd
import re
import subprocess
from openpyxl import load_workbook

# ====================================================================
# FOLLOWING IS ASSIGNMENT SETUP
# ====================================================================
# Marks components
UNSTYLE_ASSIGNMENTS_NUM = 5
ASSIGNMENTS_NUM = 10
EXAM_NUM = 1
ICLICKER_WEIGHT = 5

# helpful convert
TOTAL_ASSESSMENT = UNSTYLE_ASSIGNMENTS_NUM + ASSIGNMENTS_NUM + EXAM_NUM
OVERALL_STYLE = TOTAL_ASSESSMENT
STYLE_ASSIGNMENTS_NUM = ASSIGNMENTS_NUM - UNSTYLE_ASSIGNMENTS_NUM
MIDTERM_INDEX = TOTAL_ASSESSMENT - 1

# ====================================================================
# FOLLOWING IS ENV VARIABLES
# ====================================================================
# Get home path
HOME = os.getenv("HOME")

# Course this script is running as
# ex. CS136
COURSENAME = getpass.getuser().upper()

CURR_TERMCODE_CMD = "/u/isg/bin/termcode"
CURR_TERMCODE = subprocess.check_output(CURR_TERMCODE_CMD, shell=True).decode('utf-8').strip()

CURRTERM_CMD = "/u/isg/bin/termcode -l"
CURRTERM = subprocess.check_output(CURRTERM_CMD, shell=True).decode('utf-8').strip()

CURR_SESSION_CMD = "/u/isg/bin/termcode -l | awk '{print tolower(substr($0, 1, 1))}'"
CURR_SESSION = subprocess.check_output(CURR_SESSION_CMD, shell=True).decode('utf-8').strip()


CURR_YEAR_CMD = "/u/isg/bin/termcode -l | awk '{print substr($0, length($0)-1, length($0))}'"
CURR_YEAR = subprocess.check_output(CURR_YEAR_CMD, shell=True).decode('utf-8').strip()

# Path will be used
TERM_FOLDER = f"{CURR_TERMCODE}_{CURR_SESSION}{CURR_YEAR}"
PATH_CURRTERM = f"{HOME}/marks/current_term/"
PATH_TERM_DATA = f"{HOME}/marks/past_terms/{TERM_FOLDER}"
PATH_A0_RESULT = f"{PATH_TERM_DATA}/a0_result.txt"
PATH_CLASSLIST = f"{PATH_TERM_DATA}/classlist.csv"
PATH_CONFIG = f"{PATH_TERM_DATA}/config.csv"
PATH_EDX_MARKS = f"{PATH_TERM_DATA}/edx_marks.csv"
PATH_EXEMPTION = f"{PATH_TERM_DATA}/exemptions.csv"
PATH_GRADEBOOK = f"{PATH_TERM_DATA}/gradebook/gradebook.xlsx"
PATH_ICLICKER = f"{PATH_TERM_DATA}/clicker_result/final_grades.csv"
PATH_MARKUS_RESULT = f"{PATH_TERM_DATA}/markus_result"
PATH_MARMOSET_RESULT = f"{PATH_TERM_DATA}/marmoset_result"
PATH_MIDTERM_RESULT = f"{PATH_TERM_DATA}/midterm"
PATH_REMARK = f"{PATH_TERM_DATA}/remarks.csv"
PATH_STATS = f"{PATH_TERM_DATA}/stats.txt"

# Path of database infromation
PATH_DB_INFO = f"{HOME}/.my.cnf"

MIDTERM = True
MARMOSET = True
MARKUS = False

# ====================================================================
# Helper Functions
# ====================================================================

def assignment_setup_reader(file_path: str):
    """
    Reads the configuration file to set up assignment details, including
    project names, full marks, weights, and identifying memory questions.

    Parameters:
    - file_path (str): The path to the configuration file.

    Returns:
    - Three elements:
        - projects_info (dict): Information about each project, including
          full marks and weight.
        - memory_questions (dict): Details of memory questions, if any.

    Example:
    projects_info,
    memory_questions,
    current_assignment = assignment_setup_reader('config.csv')
    """
    print(">> Loding config")
    projects_info = {}
    memory_questions = {}
    current_assignment = 0
    with open(file_path, mode='r') as infile:
        reader = csv.DictReader(infile)
        for row in reader:
            project_name = row['project']
            if not "#" in project_name:
                if project_name == 'midterm':
                    current_assignment = project_name.capitalize()
                else:
                    current_assignment = f"Assignment {project_name[1]}"
                project_fullmark = int(row['fullMarks'])
                project_weight = int(row['weight'])
                marking_type = row['isHandMarking']
                if marking_type == '1':
                    memory_questions[project_name] = {'complete': False, 
                                                      'marmoset_path': '', 
                                                      'markus_path': ''}
                if marking_type == '2':
                    style_weight = int(row['styleWeight'])
                else:
                    style_weight = 0
                projects_info[project_name] = {'fullMark': project_fullmark, 
                                               'weight': project_weight, 
                                               'styleWeight': style_weight}
    print(f"   |> Latest assessment in config is {current_assignment}")
    return projects_info, memory_questions


def load_result_dict(classlist_path: str):
    """
    Initializes a dictionary for storing student marks based on the class list.

    Parameters:
    - classlist_path (str): The path to the class list file.

    Returns:
    - dict: A dictionary with student IDs as keys and lists of zeros for marks.

    Example:
    data_dict = load_result_dict('classlist.csv')
    """
    data_dict = {}
    with open(classlist_path, mode='r') as file:
        reader = csv.reader(file)
        data_dict['mark_status'] = [0.0 for _ in range(TOTAL_ASSESSMENT)]
        for row in reader:
            if row:
                uw_id = row[1]
                data_dict[uw_id] = [{'total': 0.0, 
                                     'assignment_part': {}} for _ in range(TOTAL_ASSESSMENT)] \
                                 + [{'total': 0.0, 
                                     'assignment_part': 
                                     {f'Assignment {i + UNSTYLE_ASSIGNMENTS_NUM}': 0.0 for i in range(STYLE_ASSIGNMENTS_NUM)}}]
    return data_dict


def load_remarks_dict(remarks_path: str):
    """
    Loads remarks for assignments from a given file, creating a dictionary to track
    these remarks by student ID and question.

    Parameters:
    - remarks_path (str): The path to the file containing remark information.

    Returns:
    - dict: A nested dictionary where each key is a question and each value is another
      dictionary mapping student IDs to their remarked grades.

    Example:
    remarks_dict = load_remarks_dict('remarks.csv')
    """
    def is_number(s):
        try:
            float(s)
            return True
        except ValueError:
            return False
        
    remarks_dict = {}
    with open(remarks_path, mode='r') as file:
        reader = csv.reader(file)
        next(reader)
        for row in reader:
            uw_id, question, new_total, markus_marks = row
            new_total = float(new_total) if is_number(new_total) else None
            markus_marks = float(markus_marks) if is_number(markus_marks) else None
            if question not in remarks_dict:
                remarks_dict[question] = {}
            remarks_dict[question][uw_id] = {'newTotal': new_total, 'markusRemark': markus_marks}
    return remarks_dict


def get_remarked_grade(questin_type: bool, project_name: str, remarks_dict: dict, uw_id: str, grades: float):
    """
    Retrieves the remarked grade for a student for a specific project, if available.

    Parameters:
    - question_type (bool): The type of question (True for marmoset, False for markus).
    - project_name (str): The name of the project.
    - remarks_dict (dict): The dictionary containing remarks.
    - uw_id (str): The student's university ID.
    - grades (float): The original grade before remark.

    Returns:
    - float: The new grade after the remark, or the original grade if no remark was found.

    Example:
    new_grade = get_remarked_grade(True, 'Project1', remarks_dict, '12345', 85.0)
    """
    if project_name in remarks_dict and uw_id in remarks_dict[project_name]:
        if questin_type:
            new_grades = remarks_dict[project_name][uw_id]['newTotal']
        else:
            new_grades = remarks_dict[project_name][uw_id]['markusRemark']
        if new_grades:
            return float(new_grades)
    return grades

# ====================================================================
# Functions
# ====================================================================

def calculate_assignments_marks(project_dict: dict, project_name: str, marks_dict: dict, remarks_dict: dict, file_path: str):
    """
    Calculates and updates the marks for assignments for each student, taking into account
    remarks and the assignment weight.

    Parameters:
    - project_dict (dict): Dictionary containing project information.
    - project_name (str): The name of the project.
    - marks_dict (dict): Dictionary to store students' marks.
    - remarks_dict (dict): Dictionary containing any remarks.
    - file_path (str): Path to the file with assignment results.

    Example:
    calculate_assignments_marks(project_info, 'Project1', marks_dict, remarks_dict, 'results.csv')
    """
    assignment_number = int(project_name[1])
    full_marks = project_dict[project_name]['fullMark']
    weight = project_dict[project_name]['weight']
    marks_dict['mark_status'][assignment_number] = 1
    with open(file_path, mode='r') as infile:
        reader = csv.reader(infile)
        for row in reader:
            uw_id = row[0]
            if uw_id in marks_dict:
                total = float(row[1])
                total = get_remarked_grade(MARMOSET, project_name, remarks_dict, uw_id, total)
                if total <= full_marks:
                    grades = total / full_marks * weight
                    marks_dict[uw_id][assignment_number]['assignment_part'][project_name] = grades
                else:
                    print(f"Please check configuration of Assignment {assignment_number} or remarks.csv")


def calculate_midterm_marks(project_dict: dict, project_name: str, marks_dict: dict, remarks_dict: dict, file_path: str):
    """
    Calculates and updates the marks for midterm for each student, taking into account
    and the assignment weight.

    Parameters:
    - project_dict (dict): Dictionary containing project information.
    - project_name (str): The name of the project.
    - marks_dict (dict): Dictionary to store students' marks.
    - file_path (str): Path to the file with midterm results.

    Example:
    calculate_midterm_marks(project_info, 'midterm', marks_dict, remarks_dict, 'results.csv')
    """
    full_marks = project_dict[project_name]['fullMark']
    weight = project_dict[project_name]['weight']
    marks_dict['mark_status'][MIDTERM_INDEX] = 1
    with open(file_path, mode='r') as infile:
        reader = csv.DictReader(infile)
        for row in reader:
            uw_id = row['Email'].split('@')[0]
            if uw_id in marks_dict:
                total = float(row['Total'])
                total = get_remarked_grade(MIDTERM, project_name, remarks_dict, uw_id, total)
                if total <= full_marks:
                    grades = total / full_marks * weight
                    marks_dict[uw_id][MIDTERM_INDEX]['assignment_part'][project_name] = grades
                else:
                    print("Please check configuration of Midterm")


def calculate_style_marks(project_dict: dict, memory_questions_dict: dict, project_name: str, marks_dict: dict, remarks_dict: dict, file_path: str):
    """
    Calculates style marks for projects, excluding memory questions, and updates the marks dictionary.

    Parameters:
    - memory_questions_dict (dict): Dictionary of memory questions to exclude from style calculation.
    - project_name (str): The name of the project.
    - marks_dict (dict): Dictionary to store students' marks.
    - remarks_dict (dict): Dictionary containing any remarks.
    - file_path (str): Path to the file with style results.

    Example:
    calculate_style_marks(memory_questions, 'Project2', marks_dict, remarks_dict, 'style_results.csv')
    """
    assignment_number = int(project_name[1])
    style_index = assignment_number + UNSTYLE_ASSIGNMENTS_NUM
    marks_dict['mark_status'][style_index] = 1
    style_weight = project_dict[project_name]['styleWeight']
    full_marks = 100.0
    if assignment_number >= UNSTYLE_ASSIGNMENTS_NUM and project_name not in memory_questions_dict:
        with open(file_path, mode='r') as infile:
            reader = csv.reader(infile)
            for row in reader:
                uw_id = row[0]
                if uw_id in marks_dict:
                    total = float(row[1])
                    total = get_remarked_grade(MARKUS, project_name, remarks_dict, uw_id, total)
                    if total <= full_marks:
                        grades = total / full_marks * style_weight
                        marks_dict[uw_id][style_index]['assignment_part'][project_name] = grades

                        grades = grades / STYLE_ASSIGNMENTS_NUM

                        marks_dict[uw_id][OVERALL_STYLE]['assignment_part'][f'Assignment {assignment_number}'] += grades
                    else:
                        print(f"Please check configuration of Assignment {assignment_number} or remarks.csv")


def calculate_memory_marks(project_dict: dict, project_name: str, project_info: dict, marks_dict: dict, remarks_dict: dict):
    """
    Calculates and updates the marks for memory-related questions in projects.

    Parameters:
    - project_dict (dict): A dictionary with project names as keys and tuples of full marks and weight as values.
    - project_name (str): The name of the current project being processed.
    - project_info (dict): A dict containing information about the project, including completion status and paths to grades.
    - marks_dict (dict): A dictionary with student IDs as keys and a list of their marks as values.
    - remarks_dict (dict): A dictionary containing any remark requests.

    Requires:
    - Correct initialization and population of project_dict, project_name, project_info, marks_dict, and remarks_dict.
    - Existence and readability of files at paths specified in project_info.

    Effects:
    - Updates `marks_dict` with calculated memory marks for the specified project.
    """
    assignment_number = int(project_name[1])
    full_marks = project_dict[project_name]['fullMark']
    weight = project_dict[project_name]['weight']
    is_complete = project_info['complete']
    marmoset_path = project_info['marmoset_path']
    markus_path = project_info['markus_path']
    marmoset_grades = {}
    if is_complete:
        with open(marmoset_path, mode='r') as infile:
            reader = csv.reader(infile)
            for row in reader:
                uw_id = row[0]
                if uw_id in marks_dict:
                    grades = float(row[1])
                    grades = get_remarked_grade(MARMOSET, project_name, remarks_dict, uw_id, grades)
                    marmoset_grades[uw_id] = grades / full_marks * weight

        with open(markus_path, mode='r') as infile:
            reader = csv.reader(infile)
            for row in reader:
                uw_id = row[0]
                if uw_id in marmoset_grades:
                    grades = float(row[1])
                    grades = get_remarked_grade(MARKUS, project_name, remarks_dict, uw_id, grades)
                    grades = marmoset_grades[uw_id] * grades / 100.0
                    marks_dict[uw_id][assignment_number]['assignment_part'][project_name] = grades
    else:
        print(f">>> {project_name} Memory snapshot missing part")
        for uw_id in marks_dict:
            marks_dict[uw_id][assignment_number] = 0.0


def finalize_marks(marks_dict: dict):
    """
    Calculates and updates the total marks for each assignment for each student in `marks_dict`. 
    It handles assignments with both on-time and extended submissions. For assignments with both 
    submission types, it compares the on-time score with the adjusted average of the on-time and 
    extended scores, updating the score to the higher of the two. The total score for each assignment 
    is then updated to reflect the sum of all relevant project scores within that assignment.

    Parameters:
    - marks_dict (dict): A nested dictionary where the outer keys are student UW IDs, and the values 
      are lists of dictionaries representing each assignment. Each assignment dictionary contains 
      an 'assignment_part' key, which itself is a dictionary of project names to scores. An 
      'extended' suffix in the project name denotes an extended submission.

    Requires:
    - `marks_dict` is properly initialized and populated with the structure mentioned.

    Effects:
    - Directly modifies `marks_dict` by updating each assignment's 'total' key to reflect the 
      sum of the project scores after comparing and possibly adjusting for extended submissions.
    - For assignments with both on-time and extended submissions for a project, the project's score 
      in 'assignment_part' is updated to the higher of the on-time score or the adjusted average.
    - Initializes the 'total' score for each assignment before summing to ensure accuracy.
    """
    for uw_id in marks_dict:
        if uw_id != 'mark_status':
            for assignment_number in range(TOTAL_ASSESSMENT + 1):
                for project in marks_dict[uw_id][assignment_number]['assignment_part']:
                    if 'extended' in project:
                        extended_total = marks_dict[uw_id][assignment_number]['assignment_part'][project]
                        project = project.split('-')[0]
                        ontime_total = marks_dict[uw_id][assignment_number]['assignment_part'][project]
                        marks_dict[uw_id][assignment_number]['assignment_part'][project] = max(ontime_total, ((ontime_total + extended_total) / 2))
                for project in marks_dict[uw_id][assignment_number]['assignment_part']:
                    if 'extended' not in project:
                        if marks_dict[uw_id][assignment_number]['total'] == 'X':
                            marks_dict[uw_id][assignment_number]['assignment_part'][project] = 0
                        else:
                            if marks_dict[uw_id][assignment_number]['assignment_part'][project] != 'X':
                                marks_dict[uw_id][assignment_number]['total'] += marks_dict[uw_id][assignment_number]['assignment_part'][project]


def set_exemptions(marks_dict: dict, exemption_file: str):
    """
    Reads an exemptions file and updates `marks_dict` to exempt certain assignments for specified students.

    Parameters:
    - marks_dict (dict): The dictionary with student IDs as keys and a list of their marks as values.
    - exemption_file (str): The file path to the exemptions file.

    Requires:
    - Correct initialization and population of `marks_dict`.
    - Existence of a valid exemptions data in `exemption_file`.

    Effects:
    - Modifies `marks_dict` to reflect exemptions based on `exemption_file`.
    """
    print(">> Processing exemptions")
    with open(exemption_file, mode='r') as infile:
        reader = csv.reader(infile)
        for row in reader:
            uw_id = row[0]
            exemp_assign = row[1]
            if uw_id in marks_dict and (len(exemp_assign) >= 2):
                if exemp_assign[1].isdigit():
                    assignment_number = int(exemp_assign[1])
                    marks_dict[uw_id][assignment_number]['total'] = 'X'
                    if assignment_number >= UNSTYLE_ASSIGNMENTS_NUM:
                        marks_dict[uw_id][assignment_number + UNSTYLE_ASSIGNMENTS_NUM]['total'] = 'X'
                        marks_dict[uw_id][OVERALL_STYLE]['assignment_part'][f'Assignment {assignment_number}'] = 'X'
                elif exemp_assign == 'MID':
                    marks_dict[uw_id][15]['total'] = 'X'


def process_marks(project_dict: dict, memory_questions_dict: dict, result_dict: dict, remarks_dict: dict, marmoset_result: str, markus_result: str, midterm_result: str):
    """
    Orchestrates the processing of marks from various sources, including memory questions, assignments, and style marks.

    Parameters:
    - project_dict (dict): Contains project names with their full marks and weights.
    - memory_questions_dict (dict): Stores information for memory question projects.
    - result_dict (dict): The main dictionary where student marks are accumulated.
    - remarks_dict (dict): Contains any remark requests.
    - marmoset_result (str): Directory path containing Marmoset results.
    - markus_result (str): Directory path containing Markus results.
    - exemption_file (str): Path to the exemptions file.

    Requires:
    - All dictionaries and paths must be correctly initialized and accessible.
    - The structure of files and directories at `marmoset_result` and `markus_result` must match expected patterns.

    Effects:
    - Processes all marks, updates `result_dict`, and handles exemptions.
    """

    print(">> Processing assignment marks")
    for file in os.listdir(marmoset_result):
        if 'project' in file:
            project_name = file.split('-')[1]
            file_path = f'{marmoset_result}/{file}'
            if project_name in project_dict:
                if 'extended' in file:
                    project_name = project_name + '-extended'
                if project_name in memory_questions_dict:
                    memory_questions_dict[project_name]['marmoset_path'] = file_path
                else:
                    calculate_assignments_marks(project_dict, project_name, result_dict, remarks_dict, file_path)

    print(">> Processing style marks")
    for project in os.listdir(markus_result):
        match = re.match(r'([a-z\d]+(?:[a-z\d]+)?)*', project)
        if match:
            project_name = match.group(0)
            file_path = f'{markus_result}/{project}'
            if project_name in project_dict:
                if project_name in memory_questions_dict:
                    memory_questions_dict[project_name]['markus_path'] = file_path
                    memory_questions_dict[project_name]['complete'] = True
                else:
                    calculate_style_marks(project_dict, memory_questions_dict, project_name, result_dict, remarks_dict, file_path)
    
    for project in memory_questions_dict:
        project_info = memory_questions_dict[project]
        calculate_memory_marks(project_dict, project, project_info, result_dict, remarks_dict)
    
    print(">> Processing midterm marks")
    for project in os.listdir(midterm_result):
        if 'midterm' in project:
            project_name = 'midterm'
            file_path = f'{midterm_result}/{project}'
            if project_name in project_dict:
                calculate_midterm_marks(project_dict, project_name, result_dict, remarks_dict, file_path)
    
    set_exemptions(result_dict, PATH_EXEMPTION)


def generate_edx_marks(edx_marks_path: str, grade_book_path: str, marks_dict: dict):
    """
    Generates a CSV file with edx marks and updates an Excel grade book with assignment marks.
    
    Parameters:
    - edx_marks_path (str): The file path where the edx marks will be saved.
    - grade_book_path (str): The file path of the grade book to be updated.
    - current_assignment (int): The current assignment number up to which marks need to be generated.
    - marks_dict (dict): A dictionary with student IDs as keys and their marks as values.
    
    Requires:
    - `edx_marks_path` and `grade_book_path` must be valid paths for writing.
    - `current_assignment` must be a non-negative integer.
    - `marks_dict` must be properly initialized and populated.
    
    Effects:
    - Creates a CSV file at `edx_marks_path` with marks up to `current_assignment`.
    - Updates or creates a grade book at `grade_book_path` with sheets for each assignment.
    """
    mark_status = marks_dict['mark_status']
    assignment_index_list = []
    with open(edx_marks_path, mode='w') as edxfile:
        writer = csv.writer(edxfile)
        header = ['']
        assignment_index = 0
        while assignment_index < ASSIGNMENTS_NUM:
            if mark_status[assignment_index]:
                header.append(f'Assignment{assignment_index}')
                assignment_index_list.append(assignment_index)
                if assignment_index >= UNSTYLE_ASSIGNMENTS_NUM:
                    style_index = assignment_index + UNSTYLE_ASSIGNMENTS_NUM
                    if mark_status[style_index]:
                        header.append(f'Assignment{assignment_index}Style')
                        assignment_index_list.append(style_index)
            assignment_index += 1
        
        if mark_status[MIDTERM_INDEX]:
            header.append('Midterm')
            assignment_index_list.append(MIDTERM_INDEX)

        del marks_dict['mark_status']

        writer.writerow(header)
        for uw_id in marks_dict:
            marks_result = [uw_id]
            for i in assignment_index_list:
                marks_result.append(round(marks_dict[uw_id][i]['total'], 5) if marks_dict[uw_id][i]['total'] != 'X' else 'X')
            writer.writerow(marks_result)
    
    print(">> Generated edx_marks.csv")

    ## grade book generater
    # Assignment marks for grade book
    for i in assignment_index_list:
        if i < ASSIGNMENTS_NUM:
            sheet_name = f'A{i}'
            rows = []
            for uw_id in marks_dict:
                student_result = marks_dict[uw_id][i]
                question_list = sorted(list(student_result['assignment_part'].keys()))
                row = [uw_id] + [student_result['assignment_part'].get(q, 0) for q in question_list] + [student_result['total']]
                rows.append(row)
            df = pd.DataFrame(rows, columns=['student'] + question_list + ['Total (100)'])

            with pd.ExcelWriter(grade_book_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                book = writer.book
                if sheet_name in book.sheetnames:
                    del book[sheet_name]
                df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    # Assignment style marks for grade book
    sheet_name = 'AStyle'
    rows = []
    for uw_id in marks_dict:
        student_result = marks_dict[uw_id][OVERALL_STYLE]
        question_list = sorted(list(student_result['assignment_part'].keys()))
        row = [uw_id] + [student_result['assignment_part'].get(q, 0) for q in question_list] + [student_result['total']]
        rows.append(row)
    df = pd.DataFrame(rows, columns=['student'] + question_list + ['Total (100)'])

    with pd.ExcelWriter(grade_book_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        book = writer.book
        if sheet_name in book.sheetnames:
            del book[sheet_name]
        df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    # Assignment iclicker marks for grade book
    sheet_name = 'iClicker'
    rows = []
    with open(PATH_ICLICKER, mode='r') as infile:
        reader = csv.reader(infile)
        for row in reader:
            uw_id = row[0]
            if uw_id in marks_dict:
                grades = float(row[1])
                row = [uw_id, grades, grades / ICLICKER_WEIGHT * 100]
                rows.append(row)
            else:
                notin.append(uw_id)
    df = pd.DataFrame(rows, columns=['student', 'Marks', 'Total (100)'])

    with pd.ExcelWriter(grade_book_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        book = writer.book
        if sheet_name in book.sheetnames:
            del book[sheet_name]
        df.to_excel(writer, sheet_name=sheet_name, index=False)
    print(">> Generated gradebook.xlsx")


def a0_pass_check(edx_marks_path: str, marks_dict: dict):
    """
    Checks for students who did not score 100 on assignment 0 and writes their email addresses to a file.

    Parameters:
    - edx_marks_path (str): The file path where the list of students who didn't score 100 on A0 will be saved.
    - marks_dict (dict): A dictionary with student IDs as keys and their marks as values.

    Requires:
    - `edx_marks_path` must be a valid path for writing.
    - `marks_dict` must be properly initialized and populated, with A0 marks included.

    Effects:
    - Writes the emails of students who scored less than 100 on A0 to the specified file.
    """
    
    with open(edx_marks_path, mode='w') as file:
        writer = csv.writer(file)
        for uw_id in marks_dict:
            a0 = round(marks_dict[uw_id][0]['total'], 2)
            if a0 < 100.0:
                writer.writerow([f'{uw_id}@uwaterloo.ca'])
    print(">> Generated a0_result.txt")

# ====================================================================
# Start of main program
# ====================================================================

def main():
    projects_info_dict, memory_questions_list = assignment_setup_reader(PATH_CONFIG)
    marks = load_result_dict(PATH_CLASSLIST)
    remarks = load_remarks_dict(PATH_REMARK)
    process_marks(projects_info_dict,
                  memory_questions_list,
                  marks,
                  remarks,
                  PATH_MARMOSET_RESULT,
                  PATH_MARKUS_RESULT,
                  PATH_MIDTERM_RESULT)
    finalize_marks(marks)
    generate_edx_marks(PATH_EDX_MARKS, PATH_GRADEBOOK, marks)
    a0_pass_check(PATH_A0_RESULT, marks)


main()
